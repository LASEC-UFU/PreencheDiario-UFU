from __future__ import annotations
import json, threading, re
from pathlib import Path
from typing import Dict, Optional
from tkinter import (
    Tk, Frame, Button, Listbox, Text, Scrollbar, Entry, LabelFrame, END, SINGLE, BOTH, LEFT, RIGHT, Y, X, TOP, BOTTOM,
    filedialog, simpledialog, messagebox, StringVar
)
from tkinter import ttk
from openpyxl import load_workbook
from selenium.webdriver.remote.webdriver import WebDriver

# project services (já existentes no seu projeto)
from services.drivers import create_driver
from services.utils import GET_URL, SEI_URL, validate_value_map, preview_text
from services.diario import fill_entries, try_click_save
from services.sei import (
    AutomacaoSEI,
    FichaDisciplina,
    STATUS_ATUALIZADA,
    STATUS_CRIADA,
    STATUS_SEM_ALTERACAO,
    formatar_relatorio_arvore,
    ler_pasta_fichas,
)

# ui & features
from ui.dialogs import ask_edit_item, choose_from_list, ask_shift_params
from features.excel_import import process_worksheet
from features.date_shift import shift_value_map


class App(Tk):
    def __init__(self):
        super().__init__()
        self.title("UFU – Diário e SEI (preenchimento visual)")
        self.geometry("1200x720")
        self.minsize(1000, 600)

        self.driver: Optional[WebDriver] = None
        self.value_map: Dict[str, str] = {}
        self.current_path: Optional[str] = None
        self.browser_var = StringVar(value="edge")
        self.sei_url_var = StringVar(value=SEI_URL)
        self.sei_folder_var = StringVar(value="Nenhuma pasta selecionada")
        self.fichas_sei: list[FichaDisciplina] = []
        self.sei_running = False

        self._build_ui()

    # ---------- UI ----------
    def _build_ui(self):
        top = Frame(self); top.pack(side=TOP, fill=X)

        ttk.Label(top, text="Browser:").pack(side=LEFT, padx=(6, 2), pady=6)
        self.cbo_browser = ttk.Combobox(
            top, textvariable=self.browser_var, state="readonly",
            values=["edge", "chrome", "firefox"], width=10
        )
        self.cbo_browser.pack(side=LEFT, padx=(0, 8), pady=6)

        self.btn_open_browser = Button(top, text="Start Navegador", command=self.on_open_browser)
        self.btn_open_browser.pack(side=LEFT, padx=4, pady=6)

        self.btn_load_json = Button(top, text="Carregar Dados", command=self.on_load_json)
        self.btn_load_json.pack(side=LEFT, padx=4, pady=6)

        self.btn_import_excel = Button(top, text="Importar Excel", command=self.on_import_excel)
        self.btn_import_excel.pack(side=LEFT, padx=4, pady=6)

        sei = LabelFrame(self, text="SEI / Fichas de componentes curriculares")
        sei.pack(side=TOP, fill=X, padx=6, pady=(0, 6))

        ttk.Label(sei, text="Endereço:").pack(side=LEFT, padx=(6, 2), pady=6)
        self.ent_sei_url = Entry(sei, textvariable=self.sei_url_var, width=30)
        self.ent_sei_url.pack(side=LEFT, fill=X, expand=True, padx=(0, 4), pady=6)
        self.btn_open_sei = Button(sei, text="Abrir SEI", command=self.on_open_sei)
        self.btn_open_sei.pack(side=LEFT, padx=4, pady=6)
        self.btn_choose_fichas = Button(
            sei, text="Selecionar pasta de fichas", command=self.on_choose_fichas
        )
        self.btn_choose_fichas.pack(side=LEFT, padx=4, pady=6)
        self.lbl_fichas = ttk.Label(sei, textvariable=self.sei_folder_var, width=28)
        self.lbl_fichas.pack(side=LEFT, padx=4, pady=6)
        self.btn_load_sei = Button(
            sei,
            text="Carregar fichas no processo",
            command=self.on_load_fichas_sei,
            state="disabled",
        )
        self.btn_load_sei.pack(side=RIGHT, padx=6, pady=6)
        self.btn_extract_sei = Button(
            sei,
            text="Extrair árvore do processo",
            command=self.on_extract_sei_tree,
            state="disabled",
        )
        self.btn_extract_sei.pack(side=RIGHT, padx=6, pady=6)

        main = Frame(self); main.pack(side=TOP, fill=BOTH, expand=True)
        left = Frame(main, width=520); left.pack(side=LEFT, fill=BOTH, expand=True)
        right = Frame(main); right.pack(side=RIGHT, fill=BOTH, expand=True)

        self.listbox = Listbox(left, selectmode=SINGLE)
        self.listbox.pack(side=TOP, fill=BOTH, expand=True, padx=6, pady=6)

        left_btns = Frame(left); left_btns.pack(side=BOTTOM, fill=X, padx=6, pady=6)
        self.btn_add = Button(left_btns, text="Insert", command=self.on_add_item); self.btn_add.pack(side=LEFT, padx=4)
        self.btn_edit = Button(left_btns, text="Edit", command=self.on_edit_item); self.btn_edit.pack(side=LEFT, padx=4)
        self.btn_remove = Button(left_btns, text="Remove", command=self.on_remove_item); self.btn_remove.pack(side=LEFT, padx=4)
        self.btn_shift = Button(left_btns, text="Ajustar datas (±)", command=self.on_shift_dates); self.btn_shift.pack(side=LEFT, padx=4)
        self.btn_save_json = Button(left_btns, text="Salvar JSON", command=self.on_save_json); self.btn_save_json.pack(side=LEFT, padx=4)
        self.btn_fill = Button(left_btns, text="Preencher diário", command=self.on_fill, state="disabled"); self.btn_fill.pack(side=RIGHT, padx=4)

        self.logs = Text(right, wrap="word", state="disabled")
        sb = Scrollbar(right, command=self.logs.yview)
        self.logs.configure(yscrollcommand=sb.set)
        self.logs.pack(side=LEFT, fill=BOTH, expand=True, padx=6, pady=6)
        sb.pack(side=RIGHT, fill=Y)
        self._log("Pronto. Use o fluxo do Diário ou abra o SEI, faça login e entre no processo desejado.")

    # ---------- Helpers ----------
    def _log(self, msg: str):
        if threading.current_thread() is not threading.main_thread():
            self.after(0, self._log, msg)
            return
        self.logs.configure(state="normal")
        self.logs.insert(END, msg + "\n")
        self.logs.see(END)
        self.logs.configure(state="disabled")

    def _log_clear(self):
        self.logs.configure(state="normal")
        self.logs.delete("1.0", END)
        self.logs.configure(state="disabled")

    def _refresh_listbox(self):
        self.listbox.delete(0, END)
        for k, v in self.value_map.items():
            self.listbox.insert(END, f"{k}: {preview_text(v)}")

    def _validate_ready(self):
        if threading.current_thread() is not threading.main_thread():
            self.after(0, self._validate_ready)
            return
        ready = (self.driver is not None) and bool(self.value_map) and not self.sei_running
        self.btn_fill.configure(state=("normal" if ready else "disabled"))
        ready_sei = (self.driver is not None) and bool(self.fichas_sei) and not self.sei_running
        self.btn_load_sei.configure(state=("normal" if ready_sei else "disabled"))
        ready_extract = (self.driver is not None) and not self.sei_running
        self.btn_extract_sei.configure(state=("normal" if ready_extract else "disabled"))

    def _set_sei_running(self, running: bool):
        self.sei_running = running
        state = "disabled" if running else "normal"
        self.btn_open_sei.configure(state=state)
        self.btn_open_browser.configure(state=state)
        self.btn_choose_fichas.configure(state=state)
        self.ent_sei_url.configure(state=state)
        self._validate_ready()

    # ---------- Ordenação por data (DD/MM/AAAA) ----------
    def _key_sort_key(self, k: str):
        """
        Converte a chave 'DD/MM/AAAA -X...' em uma tupla (ano, mês, dia, chave).
        Se a chave não casar com o padrão, manda para o final.
        """
        m = re.match(r"^\s*(\d{2})/(\d{2})/(\d{4})\s*-", k or "")
        if m:
            d, mth, y = m.groups()
            return (int(y), int(mth), int(d), k)
        return (9999, 12, 31, k)

    def _sorted_by_date(self, mapping: dict[str, str]) -> dict[str, str]:
        """Retorna um novo dict ordenado pela data ascendente."""
        return dict(sorted(mapping.items(), key=lambda kv: self._key_sort_key(kv[0])))

    # ---------- Actions ----------
    def on_open_browser(self):
        def _run():
            try:
                browser = (self.browser_var.get() or "edge").strip().lower()
                if self.driver is None:
                    self._log(f"[UI] Abrindo {browser.title()}...")
                    self.driver = create_driver(browser=browser, logger=self._log)
                self._log(f"[UI] Navegando para: {GET_URL}")
                self.driver.get(GET_URL)
                self._validate_ready()
            except Exception as e:
                self._log(f"[ERRO] Falha ao abrir navegador: {e}")
        threading.Thread(target=_run, daemon=True).start()

    def on_open_sei(self):
        url = self.sei_url_var.get().strip()
        if not re.match(r"^https?://", url, flags=re.IGNORECASE):
            messagebox.showerror(
                "Endereço do SEI",
                "Informe uma URL iniciada por http:// ou https://.",
                parent=self,
            )
            return
        browser = (self.browser_var.get() or "edge").strip().lower()

        def _run():
            try:
                if self.driver is None:
                    self._log(f"[UI] Abrindo {browser.title()}...")
                    self.driver = create_driver(browser=browser, logger=self._log)
                self._log(f"[SEI] Navegando para: {url}")
                self.driver.get(url)
                self._log("[SEI] Faça o login manualmente e abra o processo que receberá as fichas.")
                self.after(0, self._validate_ready)
            except Exception as e:
                self._log(f"[ERRO] Falha ao abrir o SEI: {e}")

        threading.Thread(target=_run, daemon=True).start()

    def on_choose_fichas(self):
        pasta = filedialog.askdirectory(
            parent=self, title="Selecione a pasta que contém as fichas HTML"
        )
        if not pasta:
            return
        try:
            fichas, erros = ler_pasta_fichas(pasta)
        except Exception as exc:
            messagebox.showerror("Fichas HTML", str(exc), parent=self)
            return
        if not fichas:
            detalhe = f"\n\n{erros[0]}" if erros else ""
            messagebox.showerror(
                "Fichas HTML",
                f"Nenhuma ficha HTML válida foi encontrada.{detalhe}",
                parent=self,
            )
            return

        self.fichas_sei = fichas
        self.sei_folder_var.set(f"{len(fichas)} fichas — {Path(pasta).name}")
        self._log(f"[SEI] Pasta selecionada: {pasta}")
        self._log(f"[SEI] {len(fichas)} fichas válidas encontradas.")
        for ficha in fichas[:10]:
            self._log(f"   - {ficha.codigo}: {ficha.nome}")
        if len(fichas) > 10:
            self._log(f"   ... e mais {len(fichas) - 10} fichas")
        for erro in erros:
            self._log(f"[AVISO] Ignorada: {erro}")
        self._validate_ready()

    def on_load_fichas_sei(self):
        if not self.driver:
            messagebox.showerror("Navegador", "Abra o SEI primeiro.", parent=self)
            return
        if not self.fichas_sei:
            messagebox.showerror("Fichas", "Selecione uma pasta com fichas HTML.", parent=self)
            return

        total = len(self.fichas_sei)
        confirmar = messagebox.askyesno(
            "Confirmar carga no SEI",
            f"Serão verificadas {total} fichas no processo atualmente aberto no SEI.\n\n"
            "Uma ficha nova será criada; uma ficha alterada será atualizada no próprio "
            "documento; uma ficha idêntica será mantida sem alterações.\n"
            "Nenhum documento será excluído.\n\n"
            "Descrição: Ficha de Componente Curricular - NOME\n"
            "Número: CÓDIGO\n"
            "Nome na Árvore: NOME\n"
            "Nível de acesso: Público\n\n"
            "O lote para no primeiro erro. Confirme que está no processo correto. Continuar?",
            parent=self,
        )
        if not confirmar:
            return

        fichas = list(self.fichas_sei)
        self._set_sei_running(True)

        def _run():
            try:
                self._log(f"[SEI] Iniciando carga de {len(fichas)} fichas...")
                automacao = AutomacaoSEI(self.driver, self._log)
                concluidas, erros = automacao.carregar_lote(fichas)
                stats = automacao.estatisticas
                resumo = (
                    f"{stats[STATUS_CRIADA]} criada(s), "
                    f"{stats[STATUS_ATUALIZADA]} atualizada(s) e "
                    f"{stats[STATUS_SEM_ALTERACAO]} sem alteração"
                )
                self._log(
                    f"[SEI] Carga encerrada: {concluidas}/{len(fichas)} verificadas — {resumo}."
                )
                if erros:
                    erro = erros[0]
                    self._log(f"[SEI] Primeira falha: {erro}")
                    self.after(
                        0,
                        lambda: messagebox.showerror(
                            "Carga interrompida",
                            f"{concluidas} ficha(s) verificada(s): {resumo}.\n\n{erro}",
                            parent=self,
                        ),
                    )
                else:
                    self.after(
                        0,
                        lambda: messagebox.showinfo(
                            "Carga concluída",
                            f"As {concluidas} fichas foram verificadas.\n\n{resumo.capitalize()}.",
                            parent=self,
                        ),
                    )
            except Exception as exc:
                self._log(f"[ERRO] Falha inesperada na carga do SEI: {exc}")
            finally:
                self.after(0, lambda: self._set_sei_running(False))

        threading.Thread(target=_run, daemon=True).start()

    def on_extract_sei_tree(self):
        if not self.driver:
            messagebox.showerror(
                "Navegador", "Abra o SEI primeiro e entre no processo desejado.", parent=self
            )
            return
        confirmar = messagebox.askyesno(
            "Extrair árvore do processo",
            "Isso vai navegar por todos os documentos do processo atualmente aberto no "
            "SEI para montar um inventário (tipo, código SEI, data e um trecho de resumo "
            "de cada um). Pode demorar em processos grandes.\n\n"
            "Os campos são lidos de forma heurística e devem ser revisados antes de usar "
            "em um parecer. Confirme que está no processo correto. Continuar?",
            parent=self,
        )
        if not confirmar:
            return

        self._set_sei_running(True)

        def _run():
            try:
                self._log("[SEI] Extraindo a árvore do processo...")
                automacao = AutomacaoSEI(self.driver, self._log)

                def progresso(indice: int, total: int, nome: str):
                    self._log(f"[SEI] {indice}/{total} — {nome}")

                documentos = automacao.extrair_arvore_processo(progresso=progresso)
                texto = formatar_relatorio_arvore(documentos)
                self._log(f"[SEI] {len(documentos)} documento(s) encontrados na árvore.")

                def _salvar():
                    caminho = filedialog.asksaveasfilename(
                        parent=self,
                        title="Salvar inventário da árvore do processo",
                        defaultextension=".txt",
                        filetypes=[("Texto", "*.txt")],
                        initialfile="relatorio_arvore_sei.txt",
                    )
                    if not caminho:
                        return
                    try:
                        Path(caminho).write_text(texto, encoding="utf-8")
                        self._log(f"[SEI] Inventário salvo em: {caminho}")
                        messagebox.showinfo(
                            "Extração concluída",
                            f"{len(documentos)} documento(s) salvos em:\n{caminho}\n\n"
                            "Revise tipo, código, data e resumo antes de usar no parecer.",
                            parent=self,
                        )
                    except Exception as exc:
                        self._log(f"[ERRO] Falha ao salvar o inventário: {exc}")
                        messagebox.showerror("Salvar inventário", str(exc), parent=self)

                self.after(0, _salvar)
            except Exception as exc:
                self._log(f"[ERRO] Falha ao extrair a árvore do processo: {exc}")
                self.after(
                    0,
                    lambda: messagebox.showerror("Extração da árvore", str(exc), parent=self),
                )
            finally:
                self.after(0, lambda: self._set_sei_running(False))

        threading.Thread(target=_run, daemon=True).start()

    def on_load_json(self):
        path = filedialog.askopenfilename(parent=self, title="Escolha dados.json", filetypes=[("JSON", "*.json")])
        if not path:
            return
        try:
            with open(path, "r", encoding="utf-8") as f:
                raw = json.load(f)
            norm, errors = validate_value_map(raw)
            self._log_clear()
            if errors:
                self._log("⚠ Erros ao validar dados.json:")
                for k, e in errors.items():
                    self._log(f" - {k}: {e}")
            self.value_map = self._sorted_by_date(norm)
            self.current_path = path
            self._refresh_listbox()
            self._log("✔ dados.json carregado. Itens:")
            for k, v in self.value_map.items():
                self._log(f" - {k}: {preview_text(v)}")
            self._validate_ready()
        except Exception as e:
            self._log_clear()
            self._log(f"[ERRO] Falha ao carregar JSON: {e}")

    def on_add_item(self):
        key = simpledialog.askstring("Nova entrada", "Informe a chave (DD/MM/AAAA -X):", parent=self)
        if not key:
            return
        val = simpledialog.askstring("Texto", "Informe o texto:", parent=self) or ""
        norm, errors = validate_value_map({key: val})
        if errors:
            msg = "\n".join(f"{k}: {e}" for k, e in errors.items())
            messagebox.showerror("Erro de validação", msg, parent=self)
            return
        nk = list(norm.keys())[0]
        if nk in self.value_map:
            messagebox.showerror("Conflito", f"A chave {nk!r} já existe.", parent=self)
            return
        self.value_map[nk] = norm[nk]
        self.value_map = self._sorted_by_date(self.value_map)
        self._refresh_listbox()
        self._log(f"[UI] Item adicionado: {nk}")
        self._validate_ready()

    def on_edit_item(self):
        sel = self.listbox.curselection()
        if not sel:
            messagebox.showinfo("Editar", "Selecione um item na lista.", parent=self)
            return
        idx = sel[0]
        old_key = list(self.value_map.keys())[idx]
        old_text = self.value_map[old_key]
        res = ask_edit_item(self, old_key, old_text)
        if not res:
            return
        new_key, new_text = res
        norm, errors = validate_value_map({new_key: new_text})
        if errors:
            msg = "\n".join(f"{k}: {e}" for k, e in errors.items())
            messagebox.showerror("Erro de validação", msg, parent=self)
            return
        new_key_norm = list(norm.keys())[0]
        new_text_norm = norm[new_key_norm]
        if new_key_norm != old_key:
            if new_key_norm in self.value_map and new_key_norm != old_key:
                messagebox.showerror("Conflito", f"A chave {new_key_norm!r} já existe.", parent=self)
                return
            del self.value_map[old_key]
        self.value_map[new_key_norm] = new_text_norm
        self.value_map = self._sorted_by_date(self.value_map)
        self._refresh_listbox()
        self._log(f"[UI] Item editado: {new_key_norm}")
        self._validate_ready()

    def on_remove_item(self):
        sel = self.listbox.curselection()
        if not sel:
            messagebox.showinfo("Remover", "Selecione um item na lista.", parent=self)
            return
        idx = sel[0]
        key = list(self.value_map.keys())[idx]
        if not messagebox.askyesno("Confirmar remoção", f"Remover a entrada '{key}'?", parent=self):
            return
        del self.value_map[key]
        self._refresh_listbox()
        self._log(f"[UI] Item removido: {key}")
        self._validate_ready()

    def on_save_json(self):
        path = filedialog.asksaveasfilename(
            parent=self, title="Salvar JSON", defaultextension=".json",
            filetypes=[("JSON", "*.json")],
            initialfile=(self.current_path or "dados.json")
        )
        if not path:
            return
        try:
            with open(path, "w", encoding="utf-8") as f:
                json.dump(self.value_map, f, ensure_ascii=False, indent=2)
            self._log(f"[UI] Salvo em: {path}")
        except Exception as e:
            self._log(f"[ERRO] Falha ao salvar JSON: {e}")

    def on_fill(self):
        if not self.driver:
            messagebox.showerror("Navegador", "Abra o navegador primeiro.", parent=self)
            return
        if not self.value_map:
            messagebox.showerror("Dados", "Carregue um dados.json válido.", parent=self)
            return
        n = len(self.value_map)
        resp = messagebox.askyesno(
            "Confirmar",
            f"Você vai preencher {n} itens na turma atual (página aberta no navegador). Continuar?",
            parent=self
        )
        if not resp:
            return

        def _run():
            try:
                self._log("Iniciando preenchimento visual...")
                ok, fail, skipped = fill_entries(self.driver, self.value_map, self._log)
                try_click_save(self.driver, self._log)
                self._log(
                    f"Preenchimento concluído: {ok} ok, {fail} não encontrado, "
                    f"{skipped} já preenchido."
                )
            except Exception as e:
                self._log(f"[ERRO] Falha no preenchimento: {e}")

        threading.Thread(target=_run, daemon=True).start()

    def on_import_excel(self):
        path = filedialog.askopenfilename(
            parent=self, title="Selecione a planilha",
            filetypes=[("Excel", "*.xlsx;*.xlsm;*.xltx;*.xltm"), ("Todos", "*.*")]
        )
        if not path:
            return
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
        except Exception as e:
            messagebox.showerror("Excel", f"Não consegui abrir o arquivo:\n{e}", parent=self)
            return

        sheetnames = list(wb.sheetnames)
        sheet = sheetnames[0] if len(sheetnames) == 1 else choose_from_list(self, "Escolha a aba", sheetnames)
        if not sheet:
            wb.close()
            return
        ws = wb[sheet]

        # Processa com o normalize/validate do projeto
        norm, stats = process_worksheet(ws, validate_value_map)
        wb.close()

        if stats.get("errors"):
            self._log("⚠ Erros ao validar itens importados:")
            for k, e in stats["errors"].items():
                self._log(f" - {k}: {e}")

        dup_on_merge = sum(1 for k in norm if k in self.value_map)
        self.value_map.update(norm)
        # >>> ORDENAR POR DATA DE VERDADE (antes estava por texto) <<<
        self.value_map = self._sorted_by_date(self.value_map)

        self._refresh_listbox()
        self._validate_ready()
        self._log("✔ Importação Excel concluída.")
        self._log(f"   Arquivo: {path}")
        self._log(f"   Aba: {sheet}")
        self._log(f"   Itens válidos: {stats.get('valid', 0)} | Ignoradas: {stats.get('skipped', 0)} | "
                  f"Sobrescritas neste lote: {stats.get('overwritten_in_lot', 0)} | "
                  f"Sobrescritas na mesclagem: {dup_on_merge}")
        self._log("   Preview:")
        for i, k in enumerate(list(norm.keys())[:5]):
            self._log(f"   - {k}: {preview_text(norm[k])}")

    def on_shift_dates(self):
        params = ask_shift_params(self)
        if not params:
            return
        unit, amount, filt = params
        new_map, stats = shift_value_map(self.value_map, unit, amount, filt)
        self.value_map = new_map
        # reforça ordenação por data após ajuste
        self.value_map = self._sorted_by_date(self.value_map)
        self._refresh_listbox()
        self._validate_ready()
        self._log(f"✔ Ajuste concluído: {stats['changed']} alteradas | inválidas: {stats['invalid']} | "
                  f"filtradas: {stats['filtered']} | sobrescritas no lote: {stats['overwritten_in_lot']}")
        self._log(f"   Unidade: {unit} | Valor: {amount:+d} | Filtro: {filt}")
        self._log("   Preview de 5 chaves após ajuste:")
        for i, k in enumerate(list(self.value_map.keys())[:5]):
            self._log(f"   - {k}")


def run():
    app = App()
    app.mainloop()
